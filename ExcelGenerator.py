import openpyxl
import openpyxl.styles.numbers
import csv
import requests
import json
import os
import re
import shutil

def createExcelDocument(dataFile, resultDir):


    resultDir = re.sub(r'.xls(x)?(?=$)', '', resultDir)
    fileName = resultDir.split('/')
    fileName = '/'.join(fileName) + '/' + fileName[-1] + '.xlsx'

    columns, file_columns, data = parse_csv(dataFile)

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(columns)

    # if not os.path.isdir(resultDir):
    #     os.makedirs(resultDir)

    columnWidth = [0] * len(columns)
    for i in range(len(columns)):
        columnWidth[i] = max(columnWidth[i], len(str(columns[i])))

    for row_id in range(2, len(data) + 2):
        for element_id in range(len(columns)):

            current_element = data[row_id - 2][element_id]
            current_column = columns[element_id]

            if current_column in file_columns and isinstance(current_element, str):
                if not current_element:
                    continue

                file_links = json.loads(current_element)

                file_folder = '/' + str(element_id) + '/' + str(row_id) + '/'
                if not os.path.isdir(resultDir + file_folder):
                    os.makedirs(resultDir + file_folder)

                ws[row_id][element_id].hyperlink = '.' + file_folder
                ws[row_id][element_id].value = "Просмотреть"

                for i in range(len(file_links)):
                    file_etension = re.search(r'.[a-zA-Z]+$', file_links[i]).group()
                    file_name = 'file_' + str(i) + file_etension

                    if re.match(r'http(s)?://.+', file_links[i]):

                        response = requests.get(file_links[i])
                        if response.status_code == 200:
                            with open(resultDir + file_folder + file_name, 'wb') as f:
                                f.write(response.content)
                        else:
                            logResponse(row_id, columns[element_id], file_links[i], resultDir)
                    else:
                        try:
                            shutil.copy2(file_links[i], resultDir + file_folder + file_name)
                        except FileNotFoundError as e:
                            logResponse(row_id, columns[element_id], file_links[i], resultDir)



                continue

            columnWidth[element_id] = max(columnWidth[element_id], len(str(current_element)))

            ws[row_id][element_id].number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[49]
            ws[row_id][element_id].value = current_element


    for i in range(len(columns)):
        ws.column_dimensions[get_letter(i)].width = (columnWidth[i] + 2) * 1.05

    wb.save(fileName)


def parse_csv(dataFile):
    with open(dataFile, encoding='utf-8') as r_file:
        file_reader = csv.reader(r_file, delimiter=";")
        counter = -1
        data = []
        for row in file_reader:
            counter += 1

            if counter == 0:
                columns = row
                continue

            if counter == 1:
                file_columns = row
                continue

            data.append(row)

    return columns, file_columns, data


def logResponse(row, columnName, link, dir):
    file = open(dir + '/' + "loadError.txt", "a")
    string = "Строка: " + str(row) + " Колонка: " + columnName + "\nНе удалось загрузить файл: " + link + "\n\n"
    file.write(string)
    file.close()

def get_letter(i):
    i += 1
    str = ''
    if i > 26:
        str = get_letter(int(i / 26) - 1)
        i = i % 26
    str += chr(64 + i)
    return str
